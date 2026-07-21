import os
from openpyxl import Workbook, load_workbook
from logger import log


class EmployeeDatabase:

    def __init__(self, employee_id, employee_name, department, salary):

        self.employee_id = employee_id
        self.employee_name = employee_name
        self.department = department
        self.salary = salary

    # -----------------------------
    # Create Workbook
    # -----------------------------

    def create_workbook(self):

        try:

            os.makedirs("Day-17/data", exist_ok=True)

            if not os.path.exists("Day-17/data/employees.xlsx"):

                workbook = Workbook()

                sheet = workbook.active

                sheet.title = "Employees"

                sheet.append(["Employee ID", "Name", "Department", "Salary"])

                workbook.save("Day-17/data/employees.xlsx")

                log("Workbook Created Successfully")

            else:

                log("Workbook Already Exists")

        except Exception as e:

            log(e)

    # -----------------------------
    # Add Employee
    # -----------------------------

    def add_employee(self):

        try:

            workbook = load_workbook("Day-17/data/employees.xlsx")

            sheet = workbook.active

            sheet.append([
                self.employee_id,
                self.employee_name,
                self.department,
                self.salary
            ])

            workbook.save("Day-17/data/employees.xlsx")

            log("Employee Added Successfully")

        except Exception as e:

            log(e)

    # -----------------------------
    # Read Employees
    # -----------------------------

    def read_employees(self):

        try:

            workbook = load_workbook("Day-17/data/employees.xlsx")

            sheet = workbook.active

            print("\nEmployee Database\n")

            for row in sheet.iter_rows(values_only=True):

                print(row)

            log("Employees Read Successfully")

        except Exception as e:

            log(e)

    # -----------------------------
    # Search Employee
    # -----------------------------

    def search_employee(self):

        try:

            workbook = load_workbook("Day-17/data/employees.xlsx")

            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):

                if row[0] == self.employee_id:

                    print("\nEmployee Found\n")

                    print(row)

                    log("Employee Found")

                    return

            print("Employee Not Found")

            log("Employee Not Found")

        except Exception as e:

            log(e)

    # -----------------------------
    # Update Department
    # -----------------------------

    def update_department(self, new_department):

        try:

            workbook = load_workbook("Day-17/data/employees.xlsx")

            sheet = workbook.active

            found = False

            for row in sheet.iter_rows():

                if row[0].value == self.employee_id:

                    row[2].value = new_department

                    found = True

            workbook.save("Day-17/data/employees.xlsx")

            if found:

                print("Department Updated")

                log("Department Updated")

            else:

                print("Employee Not Found")

                log("Employee Not Found")

        except Exception as e:

            log(e)

    # -----------------------------
    # Delete Employee
    # -----------------------------

    def delete_employee(self):

        try:

            workbook = load_workbook("Day-17/data/employees.xlsx")

            sheet = workbook.active

            found = False

            for row in range(2, sheet.max_row + 1):

                if sheet.cell(row=row, column=1).value == self.employee_id:

                    sheet.delete_rows(row)

                    found = True

                    break

            workbook.save("Day-17/data/employees.xlsx")

            if found:

                print("Employee Deleted")

                log("Employee Deleted")

            else:

                print("Employee Not Found")

                log("Employee Not Found")

        except Exception as e:

            log(e)