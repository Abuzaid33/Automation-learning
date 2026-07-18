from openpyxl import Workbook ,load_workbook
import os

os.makedirs("Day-17/data" , exist_ok=True)
workbook = Workbook()

sheet = workbook.active

sheet.title = "Employee Sheets"

sheet["A1"] = "Employee ID"
sheet["B1"] = "Name"
sheet["C1"] = "Department"
sheet["D1"] = "Salary"

sheet["A2"] = "01"
sheet["B2"] = "Abuzaid"
sheet["C2"] = "CSE"
sheet["D2"] = "25000"

sheet["A3"] = "02"
sheet["B3"] = "Basit"
sheet["C3"] = "SE"
sheet["D3"] = "30000"

sheet["A4"] = "03"
sheet["B4"] = "Ali"
sheet["C4"] = "CS"
sheet["D4"] = "16000"

sheet["A5"] = "04"
sheet["B5"] = "Rehman"
sheet["C5"] = "Networking"
sheet["D5"] = "70000"

sheet["A6"] = "05"
sheet["B6"] = "Kashif"
sheet["C6"] = "IT"
sheet["D6"] = "60000"


workbook.save("Day-17/data/employee.xlsx")


reading = load_workbook("Day-17/data/employee.xlsx")

read = reading.active

for row in read.iter_rows(values_only=True):
    print(row)